#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import openpyxl

print(os.path.abspath('.')) # 打印当前目录
xlfs = [x for x in os.listdir('.') if os.path.isfile(x)
and os.path.splitext(x)[1] == '.xlsx'] # 罗列目录内所有xlsx文件
print('需要统计',len(xlfs) , '个表格')
print (xlfs)
print(xlfs[0])

xl0 = xlfs[0]
data0 = []#复制表头数据
wb0 = openpyxl.load_workbook(filename = xl0)
ws0 = wb0.active
for i in range(1,ws0.max_column+1):
    data0.append(ws0.cell(row = 1,column = i).value)
print('表头',data0)

data1 = []#复制数据
num = len(xlfs)
for n in range(num):
    xf = xlfs[n]
    wb1 = openpyxl.load_workbook(filename = xf)
    ws1 = wb1.active
    for i in range(2,ws1.max_row + 1):
        list = []
        for j in range(1,ws1.max_column + 1):
            list.append(ws1.cell(row=i,column=j).value)
        data1.append(list)
print('数据',data1)

# # 汇总表头和数据,新建保存总表
data=[]
data.append(data0)#添加表头
for l in range(len(data1)):#添加数据
    data.append(data1[l])
wb = openpyxl.Workbook()#新建表
ws = wb.active
ws.title = '汇总'
for n_row in range(1,len(data)+1):#写入数据
    for n_col in range(1,len(data[n_row-1])+1):
        ws.cell(row=n_row,column=n_col,value=str(data[n_row-1][n_col-1]))
wb.save(filename='总表.xlsx')#保存xlsx
print ('汇总完成')
